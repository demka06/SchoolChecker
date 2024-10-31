from datetime import datetime
from re import sub, compile, search
from os.path import exists, join
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles.differential import DifferentialStyle

from pandas import DataFrame, read_excel, ExcelWriter
from loguru import logger
from bs4 import BeautifulSoup as bs, BeautifulSoup
from openpyxl.workbook import Workbook
from openpyxl.formatting.rule import ColorScale, FormatObject, ColorScaleRule, Rule
from requests import Session, Response
from requests.exceptions import ConnectionError, MissingSchema

from exm_report import EXAMPLE_REPORT

"""
- Доделать .xlsx отчет
  * Добавить легенду
  * Задать ширину колонок
  * Валидация по цветам
  * Добавить колонку "сайт активен"
- Переписать checker
  * Исправить весь пиздец...
- Проверить колонку 8, 9, 10, 11, 14, 17 по файлу ./Отчет_31_10_24_47.xlsx
- Добавить прощание
- Добавить выбор места сохранения файла
- Обернуть все в докер
"""

YES = ("да", "д", "y", "yes", "d", "у")
SUPPORTING_FILES = ("xlsx", "xls")
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 "
                  "Safari/537.36 Edg/130.0.0.0",
    "Accept-Encoding": "*",
    "Connection": "keep-alive"
}
DEFAULT_HEADER_XLSX = [
    "Краткое наименование ОООД",
    "Адрес сайта",
    "1",
    "2",
    "3",
    "4",
    "5",
    "6",
    "7",
    "8",
    "9",
    "10",
    "11",
    "12",
    "13",
    "14",
    "15",
    "16",
    "17",
    "18"
]


def check_datafile(path: str) -> bool:
    """Проверка наличия файла и расширения

    Note:
        Возвращает :obj:`True` если файл есть. :obj:`False` если файла нет
    Args:
        path (:obj:`str`): путь до файла.
    Return:
        :obj:`bool`
        """

    # logger.debug(f"{exists(path), list(filter(path.endswith, SUPPORTING_FILES))}")
    return exists(path) and list(filter(path.endswith, SUPPORTING_FILES))


def restart() -> bool:
    """Перезапуск / завершение работы программы

    Note:
        Возвращает :obj:`True` если требуется перезапуск. :obj:`False` завершение работы программы
    Return:
        :obj:`bool`:
    """

    restart_flag = input("\033[37mНачать выполнение программы сначала? (Д/Н): ").lower()
    return True if restart_flag in YES else False


def get_query(url: str) -> Response | int:
    """Отправка GET запроса, а также валидация исключений

    Note:
        Возвращает :obj:`requests.Response` при корректном выполнении запроса.
        :obj:`int` при некорректном выполнении запроса.
    Args:
        url (:obj:`str`): URL сайта
    Return:
        :obj:`requests.Response` | :obj:`int`
    """
    with Session() as session:
        logger.debug(url)

        if not url:
            logger.warning(f"Не указана ссылка на сайт: {url}")
            return 4

        if search(r"h+t+p+s?:?/+", url):
            url = sub(r"h+t+p+s?:?/+", "http://", url)
        else:
            url = "http://" + url

        try:
            with session.get(url=url, headers=HEADERS) as response:
                if response is None:
                    return 3
                return response
        except ConnectionError:
            logger.warning(f"Сайт {url} не отвечает на запросы")
            return 1
        except MissingSchema:
            return 2


def get_soup(content) -> BeautifulSoup:
    """Возвращает объект для последующего парсинга

    Return:
        :obj:`bs4.BeautifulSoup`
    """
    return bs(content, "lxml")


def xlsx_to_dict(file: str) -> list[dict]:
    """Преобразует .xlsx в tuple

    Args:
         file (:obj:`tuple`): путь до файла
    Return:
         :obj:`tuple` преобразованные данные
    """
    pd = read_excel(file, engine="openpyxl", index_col=0, skiprows=0)
    dataframe = DataFrame(pd)
    dataframe = dataframe.rename(columns={
        "Краткое наименование ОООД": "org",
        "Адрес сайта": "url"
    })
    return dataframe.to_dict(orient="records")


def create_empty_xlsx_report() -> openpyxl.workbook.workbook:
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.freeze_panes = "B2"

    ws.append(DEFAULT_HEADER_XLSX)

    return wb


def save_xlsx_report(wb: openpyxl.workbook.workbook) -> str:
    filename = f"Отчет_{datetime.now().strftime("%d_%m_%y_%S")}.xlsx"
    filepath = join('../', filename)
    wb.save(filepath)
    return filepath


def add_data_to_report(wb: openpyxl.workbook.workbook, rep, school: str, link: str):
    ws = wb.active

    ws.append([school,
               link,
               "+" if rep["osn_sved_razd"]["options"][0]["osn_sved_oo"] else "-",
               "+" if rep["osn_sved_razd"]["options"][1]["nazv_oo"] else "-",
               "+" if rep["osn_sved_razd"]["options"][2]["inf_uchred_oo"] else "-",
               "+" if rep["osn_sved_razd"]["options"][3]["geo_oo"] else "-",
               "+" if rep["osn_sved_razd"]["options"][4]["grafik_oo"] else "-",
               "+" if rep["osn_sved_razd"]["options"][5]["phones_oo"] else "-",
               "+" if rep["osn_sved_razd"]["options"][6]["mail_oo"] else "-",
               "+" if rep["osn_sved_razd"]["options"][7]["license_oo"] else "-",
               "+" if rep["osn_sved_razd"]["options"][8]["accredit_oo"] else "-",
               "+" if rep["struct_razd"]["options"][0]["nazv_podr"] else "-",
               "+" if rep["struct_razd"]["options"][1]["geo_podr"] else "-",
               "+" if rep["struct_razd"]["options"][2]["mail_podr"] else "-",
               "+" if rep["struct_razd"]["options"][3]["fio_podr"] else "-",
               "+" if rep["doc_razd"]["options"][0]["ustav_doc"] else "-",
               "+" if rep["doc_razd"]["options"][1]["raspor_doc"] else "-",
               "+" if rep["doc_razd"]["options"][2]["trudo_doc"] else "-",
               "+" if rep["doc_razd"]["options"][3]["pravila_priema_doc"] else "-",
               "+" if rep["doc_razd"]["options"][4]["rasp_doc"] else "-"
               ])


def format_xlsx_report(wb: openpyxl.workbook.workbook) -> openpyxl.workbook.workbook:
    ws = wb.active
    red_fill = PatternFill(bgColor="FFC7CE")
    green_fill = PatternFill(bgColor="008000")
    dxf_red = DifferentialStyle(fill=red_fill)
    dxf_green = DifferentialStyle(fill=green_fill)
    red_false = Rule(type="containsText", operator="containsText", text="-", dxf=dxf_red)
    green_true = Rule(type="containsText", operator="containsText", text="+", dxf=dxf_green)
    red_false.formula = ['NOT(ISERROR(SEARCH("-",C1)))']
    green_true.formula = ['NOT(ISERROR(SEARCH("+",C1)))']
    ws.conditional_formatting.add('C1:X50', red_false)
    ws.conditional_formatting.add('C1:X50', green_true)

    return wb


def change_size(wb: openpyxl.workbook.workbook) -> openpyxl.workbook.workbook:
    ws = wb.active
    cols_dict = {}
    font_size = 11
    for row in ws.rows:
        for cell in row:
            letter = cell.column_letter
            if cell.value:
                len_cell = len(str(cell.value))
                len_cell_dict = 0

                if letter in cols_dict:
                    len_cell_dict = cols_dict[letter]

                if len_cell > len_cell_dict:
                    cols_dict[letter] = len_cell
                    new_width_col = len_cell * font_size ** (font_size * 0.01)
                    ws.column_dimensions[cell.column_letter].width = new_width_col

    return wb


class Checker:
    def __init__(self, page: Response, school: str):
        self.page = page
        self.report = EXAMPLE_REPORT
        self.school = school

    def check_status(self) -> None:
        self.report["site_work"]["status"] = True if self.page.status_code == 200 else False

    def check(self):
        chapters = [i for i in self.report.keys() if i.endswith("razd")]
        print(self.page)
        main_soup = get_soup(self.page.content)

        for chapter in chapters:

            pattern_text = EXAMPLE_REPORT[chapter]["pattern"]["text"]
            elem = main_soup.select_one(f"a:contains('{pattern_text[0]}')")

            if elem is None:
                self.report[chapter]["status"] = False
                continue

            self.report[chapter]["status"] = True

            if not elem.attrs['href'].startswith(self.page.url):
                if self.page.url.endswith("/"):
                    elem.attrs['href'] = self.page.url[:-1] + elem.attrs['href']
                else:
                    elem.attrs['href'] = self.page.url + elem.attrs['href']

            self.report[chapter]["link"] = elem.attrs['href']

            slave_page = get_query(elem.attrs['href'])

            if type(slave_page) is int:
                continue

            slave_soup = get_soup(slave_page.content)

            for i in range(0, EXAMPLE_REPORT[chapter]["options"].__len__()):
                option = EXAMPLE_REPORT[chapter]["options"][i]

                keys = option.keys()
                if not keys:
                    continue

                regular = option["re"]
                if slave_soup.find_all(string=compile(regular)):
                    self.report[chapter]["options"][i][list(keys)[0]] = True

        return self.report


def core():
    while True:

        PATH_TO_FILE = input("\n\033[37m[Ввод] Введите путь до файла с адресами сайтов: ")

        if not check_datafile(PATH_TO_FILE):
            print("\033[31m[Ошибка]: ",
                  "\033[37mФайл либо отсутствует по указанному пути, либо имеет не поддерживаемое "
                  "расширение.")
            if restart():
                continue
            exit(0)

        data = xlsx_to_dict(PATH_TO_FILE)

        full_report = create_empty_xlsx_report()

        for obj in data:
            page = get_query(obj["url"])
            if type(page) is int:
                continue

            checker_obj = Checker(page, obj["org"])
            report = checker_obj.check()
            add_data_to_report(full_report, report, obj["org"], obj["url"])

        formatted_report = format_xlsx_report(full_report)
        changed_size = change_size(formatted_report)
        path_to_report = save_xlsx_report(changed_size)

        print(f"\033[33m[Сообщение] Отчет: {path_to_report}")

        if restart():
            continue
        exit(0)


core()
